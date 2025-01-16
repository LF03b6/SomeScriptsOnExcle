import openpyxl
from collections import Counter

# 处理 Excel 文件，标记重复的行
def mark_duplicates(file_path):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 获取活动工作表（默认是第一个sheet）
    
    # 读取所有 A 列的数据（从第二行开始，因为第一行一般是标题）
    a_column_values = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]

    # 使用 Counter 统计 A 列中每个值出现的次数
    counts = Counter(a_column_values)
    
    # 遍历 A 列的数据，如果某个值出现多次，标记后续行的重复值
    for row in range(2, ws.max_row + 1):  # 从第二行开始
        cell_value = ws.cell(row=row, column=1).value
        if counts[cell_value] > 1:
            # 标记后续出现的重复行背景色为红色
            for check_row in range(row + 1, ws.max_row + 1):
                check_cell_value = ws.cell(row=check_row, column=1).value
                if check_cell_value == cell_value:
                    ws.cell(row=check_row, column=1).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # 保存修改后的文件
    wb.save("output.xlsx")
    # print(f"文件已保存为 'marked_{file_path}'")

# 主函数，传入文件路径
if __name__ == '__main__':
    file_path = 'input.xlsx'  # 替换成你的文件路径
    mark_duplicates(file_path)
