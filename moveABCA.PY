import openpyxl

# 加载工作簿
wb = openpyxl.load_workbook('input.xlsx')
sheet = wb.active

# 遍历每一行，合并数据
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    cell_a = row[0]
    cell_b = row[1]
    cell_c = row[2]

    new_value = str(cell_a.value) + str(cell_b.value) + str(cell_c.value)
    cell_a.value = new_value

# 删除B列和C列
sheet.delete_cols(2, 2)

# 保存修改后的Excel文件
wb.save('output.xlsx')