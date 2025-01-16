import openpyxl

# 数字到汉字的映射字典
num_to_chinese = {
    '0': '零',
    '1': '一',
    '2': '二',
    '3': '三',
    '4': '四',
    '5': '五',
    '6': '六',
    '7': '七',
    '8': '八',
    '9': '九'
}

# 加载Excel文件
workbook = openpyxl.load_workbook('input.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    cell = row[0]
    if cell.value:
        value_str = str(cell.value)
        new_value = ""
        for char in value_str:
            if char.isdigit():
                new_value += num_to_chinese[char]
            elif char.islower():
                new_value += char.upper()
            else:
                new_value += char
        cell.value = new_value

# 保存修改后的文件
workbook.save('output.xlsx')