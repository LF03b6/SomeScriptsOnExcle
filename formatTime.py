import openpyxl
from datetime import datetime
import re

# 将日期格式化为 '2006-01-02 15:04:05' 格式
def format_date(date_str):
    try:
        # 处理 2013/10/1 或类似的日期格式
        if '/' in date_str:
            date_obj = datetime.strptime(date_str, '%Y/%m/%d')
        # 处理 2012年10月 或类似的日期格式
        elif '年' in date_str and '月' in date_str:
            date_obj = datetime.strptime(date_str, '%Y年%m年')
        else:
            # 如果格式不匹配，返回原始值
            return date_str
        
        # 返回格式化的日期时间
        return date_obj.strftime('%Y-%m-%d %H:%M:%S')
    
    except Exception as e:
        print(f"Error formatting date: {date_str}, Error: {e}")
        return date_str

# 打开 Excel 文件
def process_excel(file_path):
    # 载入 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 获取活动表（默认是第一个sheet）

    # 遍历 J 列（索引为9，从第二行开始）
    for row in range(2, ws.max_row + 1):  # 从第二行开始遍历
        cell = ws.cell(row=row, column=10)  # J 列是第10列（索引从1开始）
        original_value = cell.value
        if original_value:
            # 格式化日期
            formatted_value = format_date(str(original_value))
            cell.value = formatted_value
    
    # 保存修改后的文件
    wb.save('output.xlsx')  # 保存为新的文件，避免覆盖原文件

# 主函数，传入文件路径
if __name__ == '__main__':
    file_path = 'input.xlsx'  # 替换为你的文件路径
    process_excel(file_path)
    print(f"文件已保存为 'formatted_{file_path}'")








